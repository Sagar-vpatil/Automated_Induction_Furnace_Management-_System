from tkinter import*
import tkinter
from tkinter import ttk
from tkinter import messagebox
import os
import openpyxl
from openpyxl import Workbook, load_workbook
import shutil

frame_count = 0

class Bill_App:
    def __init__(self,root):


        self.date=StringVar()
        self.shift=StringVar()
        self.sup=StringVar()
        self.melter=StringVar()

        #Column 1

        self.IFNo=StringVar()
        self.Heattd=StringVar()
        self.Pallet=StringVar()
        self.Akg=StringVar()
        self.Akg1=StringVar()
        self.Akg2=StringVar()
        self.Akg3=StringVar()
        self.Akg4=StringVar()
        self.Fkg=StringVar()
        self.Fkg1=StringVar()
        self.Fkg2=StringVar()
        self.timing=StringVar()
        self.timing1=StringVar()
        self.timing2=StringVar()
        self.timing3=StringVar()
        self.timing4=StringVar()
        self.ts=StringVar()
        self.te=StringVar()
        
        #Column 2
        self.C2_IFNo=StringVar()
        self.C2_Heattd=StringVar()
        self.C2_Pallet=StringVar()
        self.C2_Akg=StringVar()
        self.C2_Akg1=StringVar()
        self.C2_Akg2=StringVar()
        self.C2_Akg3=StringVar()
        self.C2_Akg4=StringVar()
        self.C2_Fkg=StringVar()
        self.C2_Fkg1=StringVar()
        self.C2_Fkg2=StringVar()
        self.C2_timing=StringVar()
        self.C2_timing1=StringVar()
        self.C2_timing2=StringVar()
        self.C2_timing3=StringVar()
        self.C2_timing4=StringVar()
        self.C2_ts=StringVar()
        self.C2_te=StringVar()




        def enter_data():
            
            #accepted = accept_var.get()
            #if accepted=="Accepted":
                # Frame 1 data
                date = self.date.get()
                shift = self.shift.get()
                sup = self.sup.get()
                melter = self.melter.get()
                
                if date != "" and shift != "" and sup != "" and melter != "":
                    # Column 1 variables
                    IFNo = self.IFNo.get()
                    Heattd = self.Heattd.get()
                    Pallet = self.Pallet.get()
                    Akg = self.Akg.get()
                    Akg1 = self.Akg1.get()
                    Akg2 = self.Akg2.get()
                    Akg3 = self.Akg3.get()
                    Akg4 = self.Akg4.get()
                    Fkg = self.Fkg.get()
                    Fkg1 = self.Fkg1.get()
                    Fkg2 = self.Fkg2.get()
                    timing = self.timing.get()
                    timing1 = self.timing1.get()
                    timing2 = self.timing2.get()
                    timing3 = self.timing3.get()
                    timing4 = self.timing4.get()
                    ts = self.ts.get()
                    te = self.te.get()
                    
                    # Column 2 variables
                    C2_IFNo = self.C2_IFNo.get()
                    C2_Heattd = self.C2_Heattd.get()
                    C2_Pallet = self.C2_Pallet.get()
                    C2_Akg = self.C2_Akg.get()
                    C2_Akg1 = self.C2_Akg1.get()
                    C2_Akg2 = self.C2_Akg2.get()
                    C2_Akg3 = self.C2_Akg3.get()
                    C2_Akg4 = self.C2_Akg4.get()
                    C2_Fkg = self.C2_Fkg.get()
                    C2_Fkg1 = self.C2_Fkg1.get()
                    C2_Fkg2 = self.C2_Fkg2.get()
                    C2_timing = self.C2_timing.get()
                    C2_timing1 = self.C2_timing1.get()
                    C2_timing2 = self.C2_timing2.get()
                    C2_timing3 = self.C2_timing3.get()
                    C2_timing4 = self.C2_timing4.get()
                    C2_ts = self.C2_ts.get()
                    C2_te = self.C2_te.get()

             

                    
                    
                    print("Date: ", date, "Shift: ", shift, "Supervisors: ", sup, "Melters: ", melter)
                    print("Induction Furance No: ", IFNo, "No of Heat Till Date: ", Heattd, "Pallet No Batch: ", Pallet)
                    print("------------------------------------------")
                    
                    
    
                    wb= load_workbook("Template.xlsx")
                    ws=wb.active
                    #add image
                    img = openpyxl.drawing.image.Image('logo.png')
                    ws.add_image(img, 'C3')
                    ws['C6'].value= date
                    ws['G6'].value= shift
                    ws['K6'].value= sup
                    ws['O6'].value= melter
                    
                    #Column 1 Values
                    ws['G8'].value= IFNo
                    ws['G9'].value= Heattd
                    ws['G10'].value= Pallet
                    ws['G13'].value= Akg
                    ws['G14'].value= Akg1
                    ws['G15'].value= Akg2
                    ws['G16'].value= Akg3
                    ws['G17'].value= Akg4
                    ws['G20'].value= Fkg
                    ws['G21'].value= Fkg1
                    ws['G22'].value= Fkg2
                    ws['G26'].value= timing
                    ws['G27'].value= timing1
                    ws['G28'].value= timing2
                    ws['G30'].value= timing3
                    ws['G31'].value= timing4
                    ws['G32'].value= ts
                    ws['G34'].value= te

                    #Column 2 Values
                    ws['H8'].value= C2_IFNo
                    ws['H9'].value= C2_Heattd
                    ws['H10'].value= C2_Pallet
                    ws['H13'].value= C2_Akg
                    ws['H14'].value= C2_Akg1
                    ws['H15'].value= C2_Akg2
                    ws['H16'].value= C2_Akg3
                    ws['H17'].value= C2_Akg4
                    ws['H20'].value= C2_Fkg
                    ws['H21'].value= C2_Fkg1
                    ws['H22'].value= C2_Fkg2
                    ws['H26'].value= C2_timing
                    ws['H27'].value= C2_timing1
                    ws['H28'].value= C2_timing2
                    ws['H30'].value= C2_timing3
                    ws['H31'].value= C2_timing4
                    ws['H32'].value= C2_te
                    ws['H34'].value= C2_ts

                    
                    


                    wb.save("Template.xlsx")

                    source = "Template.xlsx"
                    destination = r"C:\Users\HP\Desktop\sp\Python\Excel_Project\Records_Excel"
                    new_name = sup + ".xlsx"

                    def copy_excel_file(source, destination, new_name):
                        #shutil.copy2(source, destination)
                        destination_path = os.path.join(destination, new_name)
                        shutil.copy2(source, destination_path)

                    copy_excel_file(source, destination, new_name)
                    tkinter.messagebox.showinfo(title="Success", message="Excel Generated Successfully")
                    

                    
                else:
                    tkinter.messagebox.showwarning(title="Error", message="Student Name, Standard and Div are required.")
            #else:
               # tkinter.messagebox.showwarning(title= "Error", message="You have not accepted the terms")


        global frame_count
        print(frame_count)
        

        #************************* Next Frame *****************************
        def next_frame():
            global frame_count
            frame_count += 1

            
            f3=LabelFrame(self.root,text=" Furance Details ",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="gold",bg=bg_color)
            f3.place(x=5,y=170,width=1525,height=90)
           

            C2_IFNo_lbl=Label(f3,text="Induction Furance No -",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0)
            C2_IFNo_Combobox=ttk.Combobox(f3,values=["A1","B1","C1"],textvariable=self.C2_IFNo,font=("times new roman",16,"bold")).grid(row=0,column=1)

            C2_Heattd_lbl=Label(f3,text="No of Heat Till Date",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=2,padx=40,pady=10,sticky="w")
            C2_Heattd_text=Entry(f3,width=10,textvariable=self.C2_Heattd,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=3,padx=10,pady=10)


            C2_Pallet_lbl=Label(f3,text="Pallet No Batch",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=4,padx=40,pady=10,sticky="w")
            C2_Pallet_text=Entry(f3,width=10,textvariable=self.C2_Pallet,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=5,padx=10,pady=10)

           #************************* Alloying *****************************
            C2_frame_al=LabelFrame(self.root,text="Alloying",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="white",bg=bg_color)
            C2_frame_al.place(x=5,y=267,width=300,height=360)

          
            C2_Akg_lbl=Label(C2_frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0,padx=10,pady=10,sticky="w")
            C2_Akg_text=Entry(C2_frame_al,width=10,textvariable=self.C2_Akg,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=1,padx=10,pady=10)

            C2_Akg1_lbl=Label(C2_frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=1,column=0,padx=10,pady=10,sticky="w")
            C2_Akg1_text=Entry(C2_frame_al,width=10,textvariable=self.C2_Akg1,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=1,column=1,padx=10,pady=10)

            C2_Akg2_lbl=Label(C2_frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=0,padx=10,pady=10,sticky="w")
            C2_Akg2_text=Entry(C2_frame_al,width=10,textvariable=self.C2_Akg2,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=1,padx=10,pady=10)

            C2_Akg3_lbl=Label(C2_frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=0,padx=10,pady=10,sticky="w")
            C2_Akg3_text=Entry(C2_frame_al,width=10,textvariable=self.C2_Akg3,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=3,column=1,padx=10,pady=10)

            C2_Akg4_lbl=Label(C2_frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=4,column=0,padx=10,pady=10,sticky="w")
            C2_Akg4_text=Entry(C2_frame_al,width=10,textvariable=self.C2_Akg4,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=4,column=1,padx=10,pady=10)
        
            #************************* Foundry *****************************
            C2_frame_fou=LabelFrame(self.root,text="Foundry",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="white",bg=bg_color)
            C2_frame_fou.place(x=310,y=267,width=320,height=360)

            C2_Charge_lbl=Label(C2_frame_fou,text="Charge WT-",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0,padx=10,pady=10,sticky="w")
            C2_Charge_Combobox=ttk.Combobox(C2_frame_fou,width=9,values=["OK","Not OK"],font=("times new roman",16,"bold")).grid(row=0,column=1)

            C2_Fkg_lbl=Label(C2_frame_fou,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=0,padx=10,pady=10,sticky="w")
            C2_Fkg_text=Entry(C2_frame_fou,width=10,textvariable=self.C2_Fkg,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=1,padx=10,pady=10)

            C2_Fkg1_lbl=Label(C2_frame_fou,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=0,padx=10,pady=10,sticky="w")
            C2_Fkg1_text=Entry(C2_frame_fou,width=10,textvariable=self.C2_Fkg1,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=3,column=1,padx=10,pady=10)

            C2_Fkg2_lbl=Label(C2_frame_fou,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=4,column=0,padx=10,pady=10,sticky="w")
            C2_Fkg2_text=Entry(C2_frame_fou,width=10,textvariable=self.C2_Fkg2,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=4,column=1,padx=10,pady=10)

            #************************* Melting Furance *****************************
            C2_frame_Mel=LabelFrame(self.root,text="Melting Furance",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="white",bg=bg_color)
            C2_frame_Mel.place(x=635,y=267,width=335,height=360)
           
            C2_timing_lbl=Label(C2_frame_Mel,text="Timing",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0,padx=10,pady=10,sticky="w")
            C2_timing_text=Entry(C2_frame_Mel,width=10,textvariable=self.C2_timing,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=1,padx=10,pady=10)

            C2_timing1_lbl=Label(C2_frame_Mel,text="Timing",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=1,column=0,padx=10,pady=10,sticky="w")
            C2_timing1_text=Entry(C2_frame_Mel,width=10,textvariable=self.C2_timing1,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=1,column=1,padx=10,pady=10)

            C2_timing2_lbl=Label(C2_frame_Mel,text="Timing",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=0,padx=10,pady=10,sticky="w")
            C2_timing2_text=Entry(C2_frame_Mel,width=10,textvariable=self.C2_timing2,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=1,padx=10,pady=10)

            C2_Melt_lbl=Label(C2_frame_Mel,text="Melt Chemistry-",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=0,padx=10,pady=10,sticky="w")
            C2_Melt_Combobox=ttk.Combobox(C2_frame_Mel,width=9,values=["OK","Not OK"],font=("times new roman",16,"bold")).grid(row=3,column=1,padx=10,pady=10)


            C2_timing3_lbl=Label(C2_frame_Mel,text="Timing",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=4,column=0,padx=10,pady=10,sticky="w")
            C2_timing3_text=Entry(C2_frame_Mel,width=10,textvariable=self.C2_timing3,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=4,column=1,padx=10,pady=10)

            C2_timing4_lbl=Label(C2_frame_Mel,text="770°C ± 20°C",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=5,column=0,padx=10,pady=10,sticky="w")
            C2_timing4_text=Entry(C2_frame_Mel,width=10,textvariable=self.C2_timing4,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=5,column=1,padx=10,pady=10)
           
           #Melting Furance 1

            C2_frame_Mel1=LabelFrame(self.root,text="",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="white",bg=bg_color)
            C2_frame_Mel1.place(x=975,y=267,width=555,height=360)

            C2_ts_lbl=Label(C2_frame_Mel1,text="Tapping Start",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=4,padx=10,pady=10,sticky="w")
            C2_ts_text=Entry(C2_frame_Mel1,width=10,textvariable=self.C2_ts,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=5,padx=10,pady=10)
        
            C2_te_lbl=Label(C2_frame_Mel1,text="Tapping End",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=4,padx=10,pady=10,sticky="w")
            C2_te_text=Entry(C2_frame_Mel1,width=10,textvariable=self.C2_te,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=3,column=5,padx=10,pady=10)
           
            C2_btn_frame1=Frame(C2_frame_Mel1,bd=7,relief=GROOVE)
            C2_btn_frame1.place(x=65,y=230,width=400,height=90)
           
            C2_Pre_btn1=Button(C2_btn_frame1,text="Previous",command=Main_Frame,bg="cadetblue",fg="white",bd=5,pady=12,width=11,font="arial 12 bold").grid(row=0,column=0,padx=40,pady=10)
            C2_next_btn1=Button(C2_btn_frame1,text="Next",bg="cadetblue",fg="white",bd=5,pady=12,width=11,font="arial 12 bold").grid(row=0,column=1,padx=10,pady=10)


        self.root=root
        self.root.geometry("1350x700+0+0")
        self.root.title("Furance Report Generator In Excel")
        bg_color="#074463"
        title=Label(self.root,text="Hitachi Astemo Braek Systems Pvt. Ltd Jalgaon",bd=12,relief=GROOVE,bg=bg_color,fg="white",font=("times new roman",30,"bold"),pady=2).pack(fill=X)
        #**************************** Student Details Frame ****************************
        f1=LabelFrame(self.root,text=" Induction Furance Melt Sheet ",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="gold",bg=bg_color)
        f1.place(x=0,y=80,relwidth=1)

        date_lbl=Label(f1,text="Date",bg=bg_color,fg="white",font=("times new roman",18,"bold")).grid(row=0,column=2,padx=20,pady=5)
        date_text=Entry(f1,width=15,textvariable=self.date,font="arial 15",bd=7,relief=SUNKEN).grid(row=0,column=3,pady=5,padx=10)

        shift_lbl=Label(f1,text="Shift",bg=bg_color,fg="white",font=("times new roman",18,"bold")).grid(row=0,column=4,padx=20,pady=5)
        shift_text=Entry(f1,width=15,textvariable=self.shift,font="arial 15",bd=7,relief=SUNKEN).grid(row=0,column=5,pady=5,padx=10)

        sup_lbl=Label(f1,text="Supervisor",bg=bg_color,fg="white",font=("times new roman",18,"bold")).grid(row=0,column=6,padx=20,pady=5)
        sup_text=Entry(f1,width=15,textvariable=self.sup,font="arial 15",bd=7,relief=SUNKEN).grid(row=0,column=7,pady=5,padx=10)

        melter_lbl=Label(f1,text="Melters",bg=bg_color,fg="white",font=("times new roman",18,"bold")).grid(row=0,column=8,padx=20,pady=5)
        melter_text=Entry(f1,width=15,textvariable=self.melter,font="arial 15",bd=7,relief=SUNKEN).grid(row=0,column=9,pady=5,padx=10)

        

        #**************************** Furance Details ****************************
        def Main_Frame():
                            
           f2=LabelFrame(self.root,text=" Furance Details ",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="gold",bg=bg_color)
           f2.place(x=5,y=170,width=1525,height=90)
           

           IFNo_lbl=Label(f2,text="Induction Furance No -",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0)
           IFNo_Combobox=ttk.Combobox(f2,values=["A1","B1","C1"],textvariable=self.IFNo,font=("times new roman",16,"bold")).grid(row=0,column=1)

           Heattd_lbl=Label(f2,text="No of Heat Till Date",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=2,padx=40,pady=10,sticky="w")
           Heattd_text=Entry(f2,width=10,textvariable=self.Heattd,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=3,padx=10,pady=10)


           Pallet_lbl=Label(f2,text="Pallet No Batch",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=4,padx=40,pady=10,sticky="w")
           Pallet_text=Entry(f2,width=10,textvariable=self.Pallet,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=5,padx=10,pady=10)

           #************************* Alloying *****************************
           frame_al=LabelFrame(self.root,text="Alloying",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="white",bg=bg_color)
           frame_al.place(x=5,y=267,width=300,height=360)

          
           Akg_lbl=Label(frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0,padx=10,pady=10,sticky="w")
           Akg_text=Entry(frame_al,width=10,textvariable=self.Akg,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=1,padx=10,pady=10)

           Akg1_lbl=Label(frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=1,column=0,padx=10,pady=10,sticky="w")
           Akg1_text=Entry(frame_al,width=10,textvariable=self.Akg1,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=1,column=1,padx=10,pady=10)

           Akg2_lbl=Label(frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=0,padx=10,pady=10,sticky="w")
           Akg2_text=Entry(frame_al,width=10,textvariable=self.Akg2,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=1,padx=10,pady=10)

           Akg3_lbl=Label(frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=0,padx=10,pady=10,sticky="w")
           Akg3_text=Entry(frame_al,width=10,textvariable=self.Akg3,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=3,column=1,padx=10,pady=10)

           Akg4_lbl=Label(frame_al,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=4,column=0,padx=10,pady=10,sticky="w")
           Akg4_text=Entry(frame_al,width=10,textvariable=self.Akg4,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=4,column=1,padx=10,pady=10)
        
           #************************* Foundry *****************************
           frame_fou=LabelFrame(self.root,text="Foundry",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="white",bg=bg_color)
           frame_fou.place(x=310,y=267,width=320,height=360)

           Charge_lbl=Label(frame_fou,text="Charge WT-",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0,padx=10,pady=10,sticky="w")
           Charge_Combobox=ttk.Combobox(frame_fou,width=9,values=["OK","Not OK"],font=("times new roman",16,"bold")).grid(row=0,column=1)

           Fkg_lbl=Label(frame_fou,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=0,padx=10,pady=10,sticky="w")
           Fkg_text=Entry(frame_fou,width=10,textvariable=self.Fkg,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=1,padx=10,pady=10)

           Fkg1_lbl=Label(frame_fou,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=0,padx=10,pady=10,sticky="w")
           Fkg1_text=Entry(frame_fou,width=10,textvariable=self.Fkg1,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=3,column=1,padx=10,pady=10)

           Fkg2_lbl=Label(frame_fou,text="Kg",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=4,column=0,padx=10,pady=10,sticky="w")
           Fkg2_text=Entry(frame_fou,width=10,textvariable=self.Fkg2,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=4,column=1,padx=10,pady=10)

           #************************* Melting Furance *****************************
           frame_Mel=LabelFrame(self.root,text="Melting Furance",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="white",bg=bg_color)
           frame_Mel.place(x=635,y=267,width=335,height=360)
           
           timing_lbl=Label(frame_Mel,text="Timing",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=0,column=0,padx=10,pady=10,sticky="w")
           timing_text=Entry(frame_Mel,width=10,textvariable=self.timing,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=0,column=1,padx=10,pady=10)

           timing1_lbl=Label(frame_Mel,text="Timing",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=1,column=0,padx=10,pady=10,sticky="w")
           timing1_text=Entry(frame_Mel,width=10,textvariable=self.timing1,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=1,column=1,padx=10,pady=10)

           timing2_lbl=Label(frame_Mel,text="Timing",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=0,padx=10,pady=10,sticky="w")
           timing2_text=Entry(frame_Mel,width=10,textvariable=self.timing2,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=1,padx=10,pady=10)

           Melt_lbl=Label(frame_Mel,text="Melt Chemistry-",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=0,padx=10,pady=10,sticky="w")
           Melt_Combobox=ttk.Combobox(frame_Mel,width=9,values=["OK","Not OK"],font=("times new roman",16,"bold")).grid(row=3,column=1,padx=10,pady=10)


           timing3_lbl=Label(frame_Mel,text="Timing",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=4,column=0,padx=10,pady=10,sticky="w")
           timing3_text=Entry(frame_Mel,width=10,textvariable=self.timing3,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=4,column=1,padx=10,pady=10)

           timing4_lbl=Label(frame_Mel,text="770°C ± 20°C",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=5,column=0,padx=10,pady=10,sticky="w")
           timing4_text=Entry(frame_Mel,width=10,textvariable=self.timing4,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=5,column=1,padx=10,pady=10)
           
           #Melting Furance 1

           frame_Mel1=LabelFrame(self.root,text="",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="white",bg=bg_color)
           frame_Mel1.place(x=975,y=267,width=555,height=360)

           ts_lbl=Label(frame_Mel1,text="Tapping Start",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=2,column=4,padx=10,pady=10,sticky="w")
           ts_text=Entry(frame_Mel1,width=10,textvariable=self.ts,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=2,column=5,padx=10,pady=10)
        
           te_lbl=Label(frame_Mel1,text="Tapping End",font=("times new roman",16,"bold"),bg=bg_color,fg="lightgreen").grid(row=3,column=4,padx=10,pady=10,sticky="w")
           te_text=Entry(frame_Mel1,width=10,textvariable=self.te,font=("times new roman",16,"bold"),bd=5,relief=SUNKEN).grid(row=3,column=5,padx=10,pady=10)
           
           btn_frame1=Frame(frame_Mel1,bd=7,relief=GROOVE)
           btn_frame1.place(x=65,y=230,width=400,height=90)
           
           sub_btn1=Button(btn_frame1,text="Previous",command= enter_data,bg="cadetblue",fg="white",bd=5,pady=12,width=11,font="arial 12 bold").grid(row=0,column=0,padx=40,pady=10)
           next_btn1=Button(btn_frame1,text="Next",command=next_frame,bg="cadetblue",fg="white",bd=5,pady=12,width=11,font="arial 12 bold").grid(row=0,column=1,padx=10,pady=10)

          
          

        Main_Frame()
       

  
         #************************* Button Frame *****************************
        f6=LabelFrame(self.root,text="Button",bd=10,relief=GROOVE,font=("times new roman",15,"bold"),fg="gold",bg=bg_color)
        f6.place(x=0,y=635,relwidth=1,height=150)

       # Accept terms
           #terms_frame = tkinter.LabelFrame(f6, text="Terms & Conditions",height=100, width=100, bd=5, relief=tkinter.GROOVE,bg=bg_color,fg="white",font=("times new roman",15,"bold"))
       
           #terms_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

           #accept_var = tkinter.StringVar(value="Not Accepted")
           #terms_check = tkinter.Checkbutton(terms_frame, text= "I accept the terms and conditions.",
                                  #variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
           #terms_check.grid(row=0, column=0)

        btn_frame=Frame(f6,bd=7,relief=GROOVE)
        btn_frame.place(x=520,width=500,height=105)

        
        sub_btn=Button(btn_frame,text="Export",command= enter_data,bg="cadetblue",fg="white",bd=5,pady=15,width=11,font="arial 15 bold").grid(row=0,column=0,padx=160,pady=5)
        

        
       

root = Tk()
obj = Bill_App(root)
root.mainloop()



       
